import sys
import csv
from pathlib import Path
import argparse
import openpyxl
UPDATE_FIELDNAMES = ["Node ID", "Node Type", "Node Title", "#REDACT"]

def get_fieldnames(template):
    with open(template, encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return reader.fieldnames

def get_update_fieldnames(template):
    fnames = UPDATE_FIELDNAMES + get_fieldnames(template)[4:]
    return fnames

class excel_sheet():
    def __init__(self, wb_path):
        self.path = Path(wb_path)
        self.wb = openpyxl.load_workbook(filename=wb_path)
        self.sheet = self.wb.active
        self.fieldnames = [x.value for x in self.sheet[1]]
        if "Node ID" in self.fieldnames or "Item ID" in self.fieldnames:
            self.update_sheet = True
        else:
            self.update_sheet = False

    def iter_records(self, fieldname_map=None):
        for row in self.sheet.iter_rows(2):
            new_fieldnames = []
            for fname in self.fieldnames:
                if fname in fieldname_map.keys():
                    new_fieldnames.append(fieldname_map[fname])
                else:
                    new_fieldnames.append(fname)
            vals = [x.value for x in row]
            yield dict(zip(new_fieldnames, vals))

    def check_fieldnames(self, template):
        base_fieldnames = get_fieldnames(template)
        for x, f in enumerate(self.fieldnames):
            try:
                pos = base_fieldnames.index(f)
                if x != pos:
                    print("Field", f, "is at column", x, "but should be at column", pos, "(should be", base_fieldnames[x], ")")
            except ValueError:
                print("Field", f, "is not in base template")
        if len(base_fieldnames) > len(additional_fieldnames):
            for f in base_fieldnames[len(additional_fieldnames):]:
                print("Field", f, "is missing from template")

    def map_fieldnames(self, base_fieldnames):
        wrong_fieldnames = [x for x in self.fieldnames if x not in base_fieldnames]
        possible_fieldnames = [x for x in base_fieldnames if x not in self.fieldnames]
        fieldname_map = {}
        for f in self.fieldnames:
            if f in base_fieldnames:
                fieldname_map[f] = f
            else:
                print(f, 'is not in template. Did you mean:')
                for count, x in enumerate(possible_fieldnames, 1):
                    print(count, ":", x)
                replacement = None
                while replacement == None:
                    i = input("->")
                    try:
                        replacement = int(i) - 1
                    except Exception as e:
                        print(i, 'is not a valid option')
                fieldname_map[f] = possible_fieldnames.pop(replacement)
        return fieldname_map

    def fix_fieldnames(self, template):
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        base_fieldnames = get_fieldnames(template)
        new_sheet.append(base_fieldnames)
        fieldname_map = self.map_fieldnames(base_fieldnames)
        for row in self.iter_records(fieldname_map):
            ordered_row = []
            for f in base_fieldnames:
                ordered_row.append(row.get(f))
            new_sheet.append(ordered_row)
        new_path = self.path.parent / (self.path.stem + '_ordered.xlsx')
        new_wb.save(new_path)

    def fix_update_fieldnames(self, template):
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active
        base_fieldnames = get_update_fieldnames(template)
        fieldname_map = self.map_fieldnames(base_fieldnames)
        base_fieldnames = [x for x in base_fieldnames if x in fieldname_map.values()]
        new_sheet.append(base_fieldnames)
        for row in self.iter_records(fieldname_map):
            ordered_row = []
            for f in base_fieldnames:
                ordered_row.append(row.get(f))
            new_sheet.append(ordered_row)
        new_path = self.path.parent / (self.path.stem + '_ordered.xlsx')
        new_wb.save(new_path)


def main(sheet, base_template):
    sheet = excel_sheet(sheet)
    if not sheet.update_sheet:
        sheet.fix_fieldnames(base_template)
    else:
        sheet.fix_update_fieldnames(base_template)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Check an input spreadsheet against a recollect template, reorder rows as required and map missing rows')
    parser.add_argument(
        'input_sheet', metavar='i', help='Excel sheet for upload')
    parser.add_argument(
        'template', metavar='t', help='ReCollect csv template for validation against')
    args = parser.parse_args()
    main(args.input_sheet, args.template)